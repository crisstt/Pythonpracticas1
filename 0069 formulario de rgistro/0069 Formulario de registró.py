from tkinter import *
from tkinter import ttk
from openpyxl import *
import mysql.connector
from tkinter import messagebox


def show_tabla_button():
    def stateBtnOp(estado):
        btn_frame_new.configure(state=estado)
        btn_frame_edit.configure(state=estado)
        btn_frame_eliminate.configure(state=estado)
        
        
        
    def stateBtnSave(estado):
        submit_btn.configure(state=estado)
        btn_frame_cancel.configure(state=estado)
        btn_frame_save.configure(state=estado)
        btn_frame_edit_save.configure(state=estado)
        
    def stateBox(estado):
        username_entry.config(state = estado)
        password_entry.config(state = estado)
        fullname_entry.config(state = estado)
        age_entry.config(state = estado)
        
    def cleanBox():
        username_entry.delete(0,END)
        password_entry.delete(0,END)
        fullname_entry.delete(0,END)
        age_entry.delete(0,END)
        
        
    def showTreeview():
        cleanGrid()
        stateBtnSave("disabled")
        stateBtnOp("normal")
        datos = consulta_username()
        for  row in datos:
            grid.insert("",END, text = row[0],values=(row[1],row[2],row[3],row[4]))
        username_entry.focus()

    def frame_new():
        selected = grid.focus()
        clave = grid.item(selected, 'text')
        id = clave
        stateBox("normal")
        stateBtnOp("disabled")
        stateBtnSave("normal")
        cleanBox()
        username_entry.focus()
        
            #print(type(row))
            #aux=aux + str(row) + '\n'

    def frame_edit():
        selected = grid.focus()
        clave = grid.item(selected, 'text')

        if clave == '':
            messagebox.showwarning("Modificar", "Debes Seleccionar un Elemento")
            #print('Debes Seleccionar un Elemento')
        else:
            stateBox('normal')
            cleanBox()
            valores = grid.item(selected, 'values')
            username_entry.insert(0,valores[0])
            password_entry.insert(0,valores[1])
            fullname_entry.insert(0,valores[2])
            age_entry.insert(0,valores[3])

            stateBtnOp("disabled")
            stateBtnSave("normal")

            username_entry.focus()
        
            
    def frame_eliminate():
        selected = grid.focus()
        clave = grid.item(selected, 'text')

        if clave == '':
            messagebox.showwarning("Eliminar", "Debes Seleccionar un Elemento")
            #print('Debes Seleccionar un Elemento')
        else:
            valores = grid.item(selected, 'values')
            data = valores[0]+"  "+valores[1]
            r = messagebox.askquestion("Eliminar","Desea Eliminar el Registro Selecionado \n {} {}". format(clave,data) )
            if r == messagebox.YES:
                n = elimina_username(clave)
                showTreeview()
                if n==1:
                    messagebox.showinfo("Eliminar", "Elemento Eliminado Correctamente")
                else:
                    messagebox.showinfo("Eliminar", "Elemento no Encontrado")
        #print(selected)
        #clave = grid.item(selected, 'text')
        #print(clave)
        #valores = grid.item(selected, 'values')
        #print(valores)
        #row = grid.item(selected)
        #print(row)

    def frame_save():
        inserta_username(username.get(),password.get(),fullname.get(),age.get())
        showTreeview()
        stateBtnSave("disabled")
        stateBtnOp("normal")
        username_entry.focus()
    def frame_saveedit():
        selected = grid.focus()
        clave = grid.item(selected, 'text')
        modifica_username(clave, username.get(),password.get(),fullname.get(),age.get())
        showTreeview()
        stateBtnSave("disabled")
        stateBtnOp("normal")
        username_entry.focus()
            
    def frame_cancel():
        stateBtnSave('disabled')
        stateBox("normal")
        stateBtnOp("normal")
        username_entry.focus()
        
    def cleanGrid():
        for item in grid.get_children():
            grid.delete(item)
    
    myWindow2 = Frame()
    myWindow2.place(x= 0, y=370,width = 93, height=380)

    btn_frame_show_info = Button(myWindow2, text = 'Show Info', command = showTreeview, bg = 'black', fg = 'white')
    btn_frame_show_info.place(x=5, y=5, width=80, height=30)
    btn_frame_new = Button(myWindow2, text = 'Nuevo', command = frame_new, bg = 'blue', fg = 'white')
    btn_frame_new.place(x=5, y=110, width=80, height=30)
    btn_frame_edit = Button(myWindow2, text = 'Modificar', command = frame_edit, bg = 'blue', fg = 'white')
    btn_frame_edit.place(x=5, y=150,width=80, height=30)
    btn_frame_eliminate = Button(myWindow2, text = 'Eliminar', command = frame_eliminate, bg = 'blue', fg = 'white')
    btn_frame_eliminate.place(x=5, y=190,width=80, height=30)
    btn_frame_cancel = Button(myWindow2, text = 'Cancelar', command = frame_cancel, bg = 'red', fg = 'white')
    btn_frame_cancel.place(x=5, y=230,width=80, height=30)
    btn_frame_save = Button(myWindow2, text = 'Guardar', command = frame_save, bg = 'dark green', fg = 'white')
    btn_frame_save.place(x=5, y=290,width=80, height=30)
    btn_frame_edit_save = Button(myWindow2, text = 'Save Edit', command = frame_saveedit, bg = 'dark green', fg = 'white')
    btn_frame_edit_save.place(x=5, y=321,width=80, height=30)
    
    myWindow3 = Frame(bg = 'yellow')
    myWindow3.place(x= 98, y=370,width = 575, height=380)
    grid = ttk.Treeview(myWindow3, columns = ('col1','col2','col3','col4'))
    grid.column('#0', width = 30, anchor = CENTER)
    grid.column('col1', width = 97, anchor = CENTER)
    grid.column('col2', width = 120,anchor = CENTER)
    grid.column('col3', width = 260,anchor = CENTER)
    grid.column('col4', width = 50,anchor = CENTER)

    grid.heading('#0', text = 'Id', anchor= CENTER)
    grid.heading('col1', text = 'Username', anchor= CENTER)
    grid.heading('col2', text = 'Pasword', anchor= CENTER)
    grid.heading('col3', text = 'Fullname', anchor= CENTER)
    grid.heading('col4', text = 'Age', anchor= CENTER)

    grid.pack(side = LEFT,fill = Y, pady= 5, padx = 2 )

    sb = Scrollbar(myWindow3,orient = VERTICAL)
    sb.pack(side=RIGHT, fill = Y)
    grid.config(yscrollcommand = sb.set)
    sb.config(command = grid.yview)
    grid['selectmode'] = 'browse'
    

    stateBox("disabled")
    stateBtnOp("normal")
    stateBtnSave("disabled")
    

    
def adding_value(adding_info):
    '''
    para poder usarlo se debe instalar desde shell
    pip install openpyxl

    wb = Workbook() permite abrebiar la llamada del archivo excel
    wb.save ('sadasdasd') guardarlo o crear el archivo en primera instancia
    '''
    #wb = Workbook()
    #wb.save ('0069 - Formulario de Usuarios.xlsx')
    '''
    se usa wb2 como variable del archiivo a trabajar
    load_workbook('asdsadasdasd') nos permite abrir o leer un archivo excel existente
    
    ws1 = wbs.create_sheet('asdasdasd', 0) nos permite crear una hoja de trabajo antes de cualquier otra hoja para trabajar en ella
    ws1.sheet_properties.tabColor = "1072FF", permite usar un color para la hoja
    ws1 = wb2.active permite trabajar en la primera hoja del documento
    wb2.active = ws1 la hoja a activar para trabajar
    
    '''
    wb2 = load_workbook ('0069 - Formulario de Usuarios.xlsx')
    #ws1 = wb2.create_sheet('USERForm',0)
    #ws1.sheet_properties.tabColor = "1072FF"
    ws1 = wb2.active
    wb2.active = ws1
    '''
    Se usa if para validar si la primera linea tiene informacion de cabecera caso que no se crea y de lo contrario se introduce solo datos
    '''
    if ws1['A1'] != 'Username':
        ws1['A1'] ='Username'
        ws1['B1'] = 'Password'
        ws1['C1'] = 'Fullname'
        ws1['D1'] = 'Age'
        ws1.append(adding_info)
    else:
        ws1.append(adding_info)
    
    '''
    sirbe para guardar el formulario de excel
    wb2 nombre de la variable que es el archivo excel
    save comando para guardar archivo
    '''
    wb2.save('0069 - Formulario de Usuarios.xlsx')
    


'''
Usamos la funcion para guardar la informacion del formu
'''
def send_data():
    adding_info =[]
    
    username_data = username.get()
    password_data = password.get()
    fullname_data = fullname.get()
    age_data =  age.get()

    #guardar para excel
    adding_info.append(username_data)
    adding_info.append(password_data)
    adding_info.append(fullname_data)
    adding_info.append(age_data)
    #funcio que guarda en excel
    adding_value(adding_info)

    #Guarda en txt sin openpyxl
    '''
    newflie = open('Formulario de Usuario.txt','a')
    newfile.write(username_data)
    newfile.write(password_data)
    newfile.write(fullname_data)
    newfile.write(age_data)
    newfile.close()
    '''
    print(username_data, '\t' , password_data, '\t', fullname_data, '\t' , fullname_data, '\t', age_data)

   




'''
--------------------------------------------------
--------------------------------------------------

'''

cnn = mysql.connector.connect(host = 'localhost', user = 'root', passwd = 'rosado1988',
                              database= 'formuser')

def str():
    pass

def consulta_username():
    cur = cnn.cursor()
    cur.execute("SELECT * FROM users")
    datos = cur.fetchall()
    cur.close()
    return datos

def buscar_username(Id):
    cur = cnn.cursor()
    sql = "SELECT * FROM users WHERE Id = {}". FORMAT(Id)
    cur.execute(sql)
    datos = cur.fetchone()
    cur.close()
    return datos

def inserta_username(Username,Password,Fullname,Age):
    cur = cnn.cursor()
    sql= '''INSERT INTO users (Username,Password,Fullname,Age)
    VALUES (' {} ',' {} ',' {} ',' {} ')'''. format (Username,Password,Fullname,Age)
    cur.execute(sql)
    n=cur.rowcount
    cnn.commit()
    cur.close()
    return n

def elimina_username(Id):
    cur = cnn.cursor()
    sql = '''DELETE FROM users WHERE Id = {} '''. format(Id)
    cur.execute(sql)
    n = cur.rowcount
    cnn.commit()
    cur.close()
    return n

def modifica_username(Id, Username,Password,Fullname,Age):
    cur = cnn.cursor()
    sql= '''UPDATE users SET Username = ' {} ', Password = ' {} ', Fullname = ' {} ', Age = ' {} ' WHERE Id = {}  '''. format(Username,Password,Fullname,Age,Id)
    cur.execute(sql)
    n=cur.rowcount
    cnn.commit()
    cur.close()
    return n


'''
--------------------------------------------------------------------------
--------------------------------------------------------------------------
'''

myWindow = Tk()
myWindow.geometry('677x750')
myWindow.title('Formulario de Registro | Python | Tkinter')
myWindow.resizable(False,False)
myWindow.config(background = '#213141')
main_title = Label(text='Formulario de Registro Python | Krizztt, Everplay', font = ('Cambria', 13), bg = '#56CD63', fg = 'white', width= '550', height = '2')
main_title.pack()

# usando clase label
username_label = Label(text = 'Username', bg = '#ffeedd')
username_label.place(x=22, y=70)
password_label = Label(text = 'Password', bg = '#ffeedd')
password_label.place(x=22, y=130)
fullname_label = Label(text = 'Fullname', bg = '#ffeedd')
fullname_label.place(x=22, y=190)
age_label = Label(text = 'Age', bg = '#ffeedd')
age_label.place(x=22, y=250)

#crear las variables usando la clase StringVar
username = StringVar()
password = StringVar()
fullname = StringVar()
age = StringVar()
# asociar el campo variable entry con stringvar
username_entry = Entry(textvariable = username, width = '40') 
password_entry = Entry(textvariable = password, width = '40', show = '*')
fullname_entry = Entry(textvariable = fullname, width = '40')
age_entry = Entry(textvariable = age, width = '40')

# usar la sentencia place para pocisionar los campos
username_entry.place(x=22, y=100) 
password_entry.place(x=22, y=160)
fullname_entry.place(x=22, y=220)
age_entry.place(x=22, y=280)

# capturados al hacer click en el boton
submit_btn = Button(myWindow, text = 'Submit Info', command = send_data, width = '30', height = '2', bg = '#00CD63')
submit_btn.place(x = 22, y = 320)
treeview_btn = Button(myWindow, text = 'Show Table', command = show_tabla_button , width = '30', height = '2', bg = '#00CD63')
treeview_btn.place(x = 350, y = 320)
myWindow.mainloop()


